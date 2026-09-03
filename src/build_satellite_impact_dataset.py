import argparse
import json
import os
from dataclasses import dataclass

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from train_dst_lstm_attention import FEATURE_COLS, _load_hourly_omni


@dataclass
class EventSource:
    name: str
    count: int


def _load_goes_exis_events(xlsx_path: str) -> list[pd.Timestamp]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    events: list[pd.Timestamp] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Expect two columns: Timestamp (text) and Datetime (datetime)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row is None or len(row) < 2:
                continue
            dt = row[1]
            if isinstance(dt, pd.Timestamp):
                events.append(dt.tz_localize(None))
            elif hasattr(dt, "year"):
                events.append(pd.Timestamp(dt).tz_localize(None))
    return events


def _load_ncei_xls_events(xls_path: str) -> list[pd.Timestamp]:
    try:
        import xlrd  # type: ignore
    except Exception:
        print(
            f"[impact] xlrd not available; cannot read {xls_path}. "
            "Install python3-xlrd or convert .xls to .xlsx.",
            flush=True,
        )
        return []

    df = pd.read_excel(xls_path, engine="xlrd")
    if df.empty:
        return []

    # Heuristic: find date + time columns and combine.
    date_cols = [c for c in df.columns if "date" in str(c).lower()]
    time_cols = [c for c in df.columns if "time" in str(c).lower()]

    events: list[pd.Timestamp] = []
    if date_cols:
        date_col = date_cols[0]
        if time_cols:
            time_col = time_cols[0]
            for d, t in zip(df[date_col], df[time_col]):
                if pd.isna(d):
                    continue
                try:
                    date = pd.to_datetime(d, errors="coerce")
                    if pd.isna(date):
                        continue
                    if pd.isna(t):
                        events.append(date)
                        continue
                    time = pd.to_datetime(t, errors="coerce")
                    if pd.isna(time):
                        # HHMM style integer
                        try:
                            hhmm = int(t)
                            hour = hhmm // 100
                            minute = hhmm % 100
                            events.append(
                                pd.Timestamp(
                                    year=date.year,
                                    month=date.month,
                                    day=date.day,
                                    hour=hour,
                                    minute=minute,
                                )
                            )
                        except Exception:
                            events.append(date)
                    else:
                        events.append(
                            pd.Timestamp(
                                year=date.year,
                                month=date.month,
                                day=date.day,
                                hour=time.hour,
                                minute=time.minute,
                                second=time.second,
                            )
                        )
                except Exception:
                    continue
        else:
            dates = pd.to_datetime(df[date_col], errors="coerce")
            for d in dates.dropna():
                events.append(pd.Timestamp(d))
    else:
        # Fallback: parse any column as datetime if possible.
        for col in df.columns:
            parsed = pd.to_datetime(df[col], errors="coerce")
            if parsed.notna().sum() > 10:
                for d in parsed.dropna():
                    events.append(pd.Timestamp(d))
                break

    return events


def _label_events(index: pd.DatetimeIndex, events: list[pd.Timestamp], horizon_hours: int) -> np.ndarray:
    if not events:
        return np.zeros(len(index), dtype=np.int8)

    idx = index.values.astype("datetime64[ns]")
    labels = np.zeros(len(idx), dtype=np.int8)
    horizon = np.timedelta64(horizon_hours, "h")

    for ev in events:
        ev_ns = np.datetime64(pd.Timestamp(ev).to_datetime64())
        start_ns = ev_ns - horizon
        start = np.searchsorted(idx, start_ns, side="left")
        end = np.searchsorted(idx, ev_ns, side="right")
        if start < end:
            labels[start:end] = 1
    return labels


def build_dataset(
    omni_csv: str,
    impact_dir: str,
    out_csv: str,
    horizon_hours: int,
):
    print("[impact] loading solar wind CSV...", flush=True)
    df = _load_hourly_omni(omni_csv)

    events: list[pd.Timestamp] = []
    sources: list[EventSource] = []

    goes_path = os.path.join(impact_dir, "g16_g17_exs_spw.xlsx")
    if os.path.exists(goes_path):
        go_events = _load_goes_exis_events(goes_path)
        events.extend(go_events)
        sources.append(EventSource("goes_exis", len(go_events)))

    for fname in ("anom5j.xls", "tdrs5j.xls"):
        path = os.path.join(impact_dir, fname)
        if os.path.exists(path):
            xls_events = _load_ncei_xls_events(path)
            events.extend(xls_events)
            sources.append(EventSource(fname, len(xls_events)))

    events = [pd.Timestamp(e).tz_localize(None) for e in events if pd.notna(e)]
    events = sorted(set(events))

    print("[impact] event sources:", flush=True)
    for src in sources:
        print(f"  - {src.name}: {src.count} events", flush=True)
    print(f"[impact] total events={len(events)}", flush=True)

    labels = _label_events(df.index, events, horizon_hours)
    df = df.copy()
    df["sat_impact_next_{:d}h".format(horizon_hours)] = labels

    os.makedirs(os.path.dirname(out_csv), exist_ok=True)
    df.to_csv(out_csv, index_label="time")
    print(f"[impact] wrote dataset -> {out_csv}", flush=True)

    meta = {
        "horizon_hours": horizon_hours,
        "sources": [src.__dict__ for src in sources],
        "total_events": len(events),
        "features": FEATURE_COLS,
    }
    meta_path = os.path.splitext(out_csv)[0] + "_meta.json"
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2)
    print(f"[impact] wrote metadata -> {meta_path}", flush=True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build satellite impact dataset from anomaly tables + solar wind")
    parser.add_argument("--omni-csv", default="data/processed/omni.csv")
    parser.add_argument("--impact-dir", default="data/impact/ncei")
    parser.add_argument("--out-csv", default="data/processed/sat_impact_dataset.csv")
    parser.add_argument("--horizon-hours", type=int, default=6)
    args = parser.parse_args()

    build_dataset(
        omni_csv=args.omni_csv,
        impact_dir=args.impact_dir,
        out_csv=args.out_csv,
        horizon_hours=args.horizon_hours,
    )
